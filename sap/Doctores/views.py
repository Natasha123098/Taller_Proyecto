from django.forms import modelform_factory
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
from django.template import loader
from openpyxl.workbook import Workbook

from Doctores.forms import DoctorFormulario
from Doctores.models import Doctor

#DoctorFormulario=modelform_factory(Doctor, exclude=[])

# Create your views here.
def Nuevo_Doctor(request):
    pagina = loader.get_template('Nuevo_Doctor.html')
    if request.method == 'GET':
        formulario= DoctorFormulario
    elif request.method == 'POST':
        formulario= DoctorFormulario(request.POST)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    datos = {'formulario': formulario}
    return HttpResponse(pagina.render(datos, request))

def Editar_Doctor(request, idDoctor):
    pagina= loader.get_template('Editar_Doctor.html')
    #Doctor= Doctor.objects.get(pk=idDoctor)
    doctor= get_object_or_404(Doctor, pk=idDoctor)
    if request.method == 'GET':
        formulario= DoctorFormulario(instance=doctor)
    elif request.method == 'POST':
        formulario= DoctorFormulario(request.POST, instance=doctor)
        if formulario.is_valid():
            formulario.save()
            return redirect('inicio')
    mensaje= {'formulario': formulario}
    return HttpResponse(pagina.render(mensaje, request))

def Ver_Doctor(request, idDoctor):
    pagina = loader.get_template('Ver_Doctor.html')
    # Doctor= Doctor.objects.get(pk=idDoctor)
    doctor = get_object_or_404(Doctor, pk=idDoctor)
    mensaje = {'doctor': doctor}
    return HttpResponse(pagina.render(mensaje, request))

def Eliminar_Doctor(request, idDoctor):

    doctor = get_object_or_404(Doctor, pk=idDoctor)
    if doctor:
        doctor.delete()
        return redirect('inicio')

def Registro_Doctor(request):
    # Obtenemos todas las personas de nuestra base de datos
    #doctores = Doctor.objects.all()
    doctores = Doctor.objects.order_by('apellido')
    # Creamos el libro de trabajo
    wb = Workbook()
    # Definimos como nuestra hoja de trabajo, la hoja activa, por defecto la primera del libro worksheet
    ws = wb.active
    # En la celda B1 ponemos el texto 'REPORTE DE PERSONAS'
    ws['B1'] = 'REPORTE DE DOCTORES'
    # Juntamos las celdas desde la B1 hasta la E1, formando una sola celda
    ws.merge_cells('B1:E1')
    # Creamos los encabezados desde la celda B3 hasta la E3
    ws['B3'] = 'NOMBRE'
    ws['C3'] = 'APELLIDO'
    ws['D3'] = 'EMAIL'
    ws['E3'] = 'ESPECIALIDAD'
    cont = 4
    # Recorremos el conjunto de personas y vamos escribiendo cada uno de los datos en las celdas
    for doctor in doctores:
        ws.cell(row=cont, column=2).value = doctor.nombre
        ws.cell(row=cont, column=3).value = doctor.apellido
        ws.cell(row=cont, column=4).value = doctor.email
        ws.cell(row=cont, column=5).value = doctor.especialidad
        cont = cont + 1
    # Establecemos el nombre del archivo
    nombre_archivo = "ReporteDoctoresExcel.xlsx"
    # Definimos que el tipo de respuesta a devolver es un archivo de microsoft excel
    response = HttpResponse(content_type="application/ms-excel")
    contenido = "attachment; filename={0}".format(nombre_archivo)
    response["Content-Disposition"] = contenido
    wb.save(response)
    return response
